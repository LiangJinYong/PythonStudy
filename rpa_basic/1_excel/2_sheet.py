from tkinter import W
from openpyxl import Workbook
wb = Workbook()
# wb.active
ws = wb.create_sheet() # 새로운 Sheet를 기본 이름으로 생성
ws.title = "MySheet" # Sheet 이름 변경
ws.sheet_properties.tabColor = "689a7b" # RGB 형태로 값을 넣어주면 탭 생상 변경

# Sheet, MySheet, YourSheet
ws1 = wb.create_sheet("YourSheet") # 주어진 이름으로 Sheet 생성
ws2 = wb.create_sheet("NewSheet", 2) # 2번째 index 에 Sheet 생성

new_ws = wb["NewSheet"] # Dictionary 형태로 Sheet 에 접근

# Sheet 복사
new_ws["A1"] = "Test"
target_ws = wb.copy_worksheet(new_ws)
target_ws.title = "Copied Sheet"

print(wb.sheetnames) # 모든 Sheet 이름 확인

wb.save("sample.xlsx")